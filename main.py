import glob
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
from collections import deque
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import os
import itertools

# 한글 폰트 설정 (Windows 예시)
plt.rc('font', family='Malgun Gothic')
plt.rcParams['axes.unicode_minus'] = False

# NumPy 1.20 이상 대응
if not hasattr(np, 'float'):
    np.float = float

###############################################################################
# (1) 엑셀 로드 및 격자 구성
###############################################################################
def load_grid(filename):
    """
    엑셀 파일에서 각 셀의 값을 읽어 'START', 'Corridor', 'warehouse'로 분류합니다.
      - 셀 값이 "START"이면 시작/종료 지점으로 지정합니다.
      - 값의 첫 글자가 대문자이면 "Corridor" (복도, 이동 가능)
      - 값의 첫 글자가 소문자이면 "warehouse" (창고, 이동 불가능)
    """
    if not os.path.exists(filename):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filename}")
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.active
    grid = {}
    start_cells = []
    max_row = ws.max_row
    max_col = ws.max_col = ws.max_column  # ensure we use max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            coord = cell.coordinate  # 예: A1 (항상 대문자)
            val = cell.value
            if val is not None and isinstance(val, str):
                val_str = val.strip()
                if val_str == "START":
                    cell_type = "START"
                elif val_str and val_str[0].isupper():
                    cell_type = "Corridor"
                elif val_str and val_str[0].islower():
                    cell_type = "warehouse"
                else:
                    cell_type = "Corridor"
            else:
                cell_type = "Corridor"
            grid[(cell.row, cell.column)] = {
                'coord': coord,
                'type': cell_type,
                'value': val
            }
            if cell_type == "START":
                start_cells.append((cell.row, cell.column))
    return grid, start_cells, max_row, max_col

###############################################################################
# (2) 좌표 변환 함수
###############################################################################
def coordinate_to_index(coord_str):
    """
    "A1" → (1,1), "C35" → (35,3) 등으로 변환합니다.
    """
    letter, number = coordinate_from_string(coord_str)
    col = column_index_from_string(letter)
    return (number, col)

###############################################################################
# (3) BFS (Corridor/START만 이동)
###############################################################################
def get_neighbors(cell, grid):
    """
    셀 (r, c)에서 상하좌우 인접 셀 중 'Corridor' 또는 'START'인 셀만 반환합니다.
    """
    r, c = cell
    neighbors = []
    for dr, dc in [(1,0), (-1,0), (0,1), (0,-1)]:
        nxt = (r+dr, c+dc)
        if nxt in grid and grid[nxt]['type'] in ["Corridor", "START"]:
            neighbors.append(nxt)
    return neighbors

def bfs(start, goal, grid):
    """
    start에서 goal까지 'Corridor'와 'START'를 통해 최단 경로를 BFS로 탐색합니다.
    경로를 (r, c) 좌표 리스트로 반환하며, 없으면 None을 반환합니다.
    """
    queue = deque([start])
    came_from = {start: None}
    while queue:
        current = queue.popleft()
        if current == goal:
            path = []
            while current is not None:
                path.append(current)
                current = came_from[current]
            path.reverse()
            return path
        for nxt in get_neighbors(current, grid):
            if nxt not in came_from:
                came_from[nxt] = current
                queue.append(nxt)
    return None

###############################################################################
# (4) 창고의 인접 복도 동적 찾기
###############################################################################
def find_warehouse_cell(warehouse_str, grid):
    """
    grid 내에서 'value'가 warehouse_str와 일치하고, 'warehouse'로 분류된 셀의 좌표를 반환합니다.
    """
    for (r, c), info in grid.items():
        if info['value'] is not None and str(info['value']).strip() == warehouse_str and info['type'] == "warehouse":
            return (r, c)
    return None

def find_adjacent_corridor(warehouse_cell, grid):
    """
    warehouse_cell(창고)의 왼쪽 또는 오른쪽(수평)에서 'Corridor' 또는 'START'인 셀을 찾아 반환합니다.
    왼쪽 우선이며, 없으면 오른쪽을 반환합니다.
    """
    if not warehouse_cell:
        return None
    r, c = warehouse_cell
    left = (r, c-1)
    if left in grid and grid[left]['type'] in ["Corridor", "START"]:
        return left
    right = (r, c+1)
    if right in grid and grid[right]['type'] in ["Corridor", "START"]:
        return right
    return None

###############################################################################
# (5) TSP를 위한 페어와이즈 거리 계산 및 최적 경로 (모든 순열 검사)
###############################################################################
def compute_all_pairwise(nodes, grid):
    """
    nodes: 노드 좌표 리스트.
    반환: dict[(i, j)] = (distance, path)
    """
    pairwise = {}
    n = len(nodes)
    for i in range(n):
        for j in range(i+1, n):
            path = bfs(nodes[i], nodes[j], grid)
            if path is None:
                dist = float('inf')
            else:
                dist = len(path)
            pairwise[(i, j)] = (dist, path)
            pairwise[(j, i)] = (dist, list(reversed(path)) if path is not None else None)
    return pairwise

def solve_tsp(nodes, pairwise):
    """
    nodes: 전체 노드 리스트 (nodes[0]는 START)
    반환: best_order (노드 인덱스 순서, START로 시작, 종료), best_cost
    """
    n = len(nodes)
    if n == 1:
        return [0, 0], 0

    intermediate = list(range(1, n))
    best_cost = float('inf')
    best_order = None

    for perm in itertools.permutations(intermediate):
        order = [0] + list(perm) + [0]
        cost = 0
        valid = True
        for i in range(len(order)-1):
            d, _ = pairwise.get((order[i], order[i+1]), (float('inf'), None))
            if d == float('inf'):
                valid = False
                break
            cost += d
        if valid and cost < best_cost:
            best_cost = cost
            best_order = order

    return best_order, best_cost

def reconstruct_full_route(order, nodes, pairwise):
    """
    order: 노드 인덱스 순서 (예: [0, 3, 1, 2, 0])
    nodes: 전체 노드 좌표 리스트
    반환: full_route (전체 (r, c) 리스트)
    """
    full_route = []
    for i in range(len(order)-1):
        _, path = pairwise[(order[i], order[i+1])]
        if path is None:
            return None
        if i > 0:
            path = path[1:]
        full_route.extend(path)
    return full_route

###############################################################################
# (6) 그룹화: 인접한 창고 셀들을 묶어 그룹별 중심과 라벨 계산
###############################################################################
def group_warehouses(grid):
    """
    grid 내의 'warehouse' 셀들을, 값의 첫 글자(예: 'a')가 같은 인접 그룹으로 묶어,
    각 그룹의 중심(centroid) 좌표와 라벨(예: "a동")을 반환합니다.
    반환: list of ((centroid_row, centroid_col), label)
    """
    visited = set()
    groups = []

    def dfs(cell, group):
        stack = [cell]
        while stack:
            cur = stack.pop()
            if cur in visited:
                continue
            visited.add(cur)
            group.append(cur)
            r, c = cur
            for dr, dc in [(1,0), (-1,0), (0,1), (0,-1)]:
                nxt = (r+dr, c+dc)
                if nxt in grid and grid[nxt]['type'] == "warehouse":
                    if str(grid[nxt]['value']).strip()[0].lower() == str(grid[cell]['value']).strip()[0].lower():
                        stack.append(nxt)
    for cell in grid:
        if cell in visited:
            continue
        if grid[cell]['type'] == "warehouse" and grid[cell]['value'] is not None:
            group = []
            dfs(cell, group)
            if group:
                avg_r = sum(r for r, _ in group) / len(group)
                avg_c = sum(c for _, c in group) / len(group)
                label = str(grid[cell]['value']).strip()[0].lower() + "동"
                groups.append(((avg_r, avg_c), label))
    return groups

###############################################################################
# (7) draw_route_segments_with_colors: 세그먼트별로 서로 다른 색으로 경로 그리기
###############################################################################
def draw_route_segments_with_colors(segments):
    """
    segments: 각 세그먼트가 (r, c) 좌표 리스트인 리스트.
    각 세그먼트를 서로 다른 색(예: tab10 컬러맵)으로 그립니다.
    각 세그먼트 중, 전체 구간의 약 25% 지점에서 진행 방향 화살표를 표시합니다.
    """
    cmap = plt.get_cmap('tab10')
    for i, seg in enumerate(segments):
        color = cmap(i % 10)
        xs = [pt[1] - 0.5 for pt in seg]
        ys = [pt[0] - 0.5 for pt in seg]
        plt.plot(xs, ys, color=color, linewidth=2)
        if len(seg) > 2:
            arrow_idx = max(1, len(seg) // 4)  # 전체 구간의 25% 지점
            if arrow_idx < len(seg):
                x0, y0 = seg[arrow_idx - 1][1] - 0.5, seg[arrow_idx - 1][0] - 0.5
                x1, y1 = seg[arrow_idx][1] - 0.5, seg[arrow_idx][0] - 0.5
                plt.annotate("", xy=(x1, y1), xytext=(x0, y0),
                             arrowprops=dict(arrowstyle="->", color=color, lw=2))

###############################################################################
# (8) highlight_input_warehouses: 입력받은 창고 셀 강조 (검은 테두리와 라벨)
###############################################################################
def highlight_input_warehouses(warehouse_inputs, grid):
    """
    warehouse_inputs: 사용자 입력받은 창고(소문자) 목록 (예: ["a12"])
    grid: 전체 셀 정보 딕셔너리.
    해당 창고 셀에 검은색 사각형 테두리를 그리고, 내부에 창고 이름을 표시합니다.
    """
    ax = plt.gca()
    for w in warehouse_inputs:
        if not w or not w[0].islower():
            continue
        wh_cell = find_warehouse_cell(w, grid)
        if wh_cell:
            r, c = wh_cell
            rect = plt.Rectangle((c - 1, r - 1), 1, 1, fill=False,
                                 edgecolor='black', linewidth=2)
            ax.add_patch(rect)
            ax.text(c - 0.5, r - 0.5, w, ha='center', va='center',
                    color='black', fontsize=12, fontweight='bold')

###############################################################################
# (9) plot_grid_and_path_segments: 최종 시각화 (세그먼트별 색, 배경, 창고 그룹, 창고 강조)
###############################################################################
def plot_grid_and_path_segments(grid, segments, max_row, max_col, tsp_nodes=None, warehouse_inputs=None):
    """
    grid: 격자 정보
    segments: 각 세그먼트 (목적지까지의 경로)를 포함하는 리스트 (각 세그먼트는 (r,c) 좌표 리스트)
    tsp_nodes: (옵션) TSP 노드 좌표 목록 (파란 점으로 표시)
    warehouse_inputs: (옵션) 사용자 입력 창고 목록 (검은 테두리로 강조)
    """
    # 배경 이미지: 복도는 옅은 흰색, 창고는 옅은 회색, START는 노랑
    image = np.zeros((max_row, max_col, 3))
    for (r, c), info in grid.items():
        if info['type'] == "Corridor":
            image[r-1, c-1] = [0.95, 0.95, 0.95]
        elif info['type'] == "warehouse":
            image[r-1, c-1] = [0.8, 0.8, 0.8]
        elif info['type'] == "START":
            image[r-1, c-1] = [1, 1, 0]
        else:
            image[r-1, c-1] = [0.95, 0.95, 0.95]

    plt.figure(figsize=(max_col/2, max_row/2))
    plt.imshow(image, origin='upper', extent=[0, max_col, max_row, 0])

    # 창고 그룹 텍스트
    groups = group_warehouses(grid)
    for (centroid, label) in groups:
        r_cent, c_cent = centroid
        plt.text(c_cent - 0.5, r_cent - 0.5, label,
                 ha='center', va='center', rotation=90,
                 color='black', fontsize=14, fontweight='bold')

    # 경로 세그먼트별 색상으로 그리기
    if segments:
        draw_route_segments_with_colors(segments)

    # TSP 노드(경유지) 파란 점 표시 (옵션)
    if tsp_nodes:
        for (r, c) in tsp_nodes:
            plt.plot(c - 0.5, r - 0.5, 'o', color='blue', markersize=8)

    # 입력된 창고(소문자) 강조
    if warehouse_inputs:
        highlight_input_warehouses(warehouse_inputs, grid)

    plt.title("최단경로 네비게이션")
    plt.axis('off')
    plt.show()

###############################################################################
# (10) 메인 함수 (파일 로드, 경유지 입력 반복, TSP 최단경로 계산)
###############################################################################
def main():
    # 1) 엑셀 파일 로드
    filename = input("엑셀 파일의 경로를 입력하세요: ").strip()
    if os.path.isdir(filename):
        print("폴더를 입력하셨습니다. 폴더 내 Excel 파일 목록:")
        excel_files = glob.glob(os.path.join(filename, "*.xlsx"))
        if not excel_files:
            print("폴더 내에 Excel 파일이 없습니다.")
            return
        for idx, file in enumerate(excel_files):
            print(f"{idx+1}: {file}")
        choice = input("사용할 Excel 파일 번호를 선택하세요: ").strip()
        try:
            index = int(choice) - 1
            filename = excel_files[index]
        except Exception as e:
            print("잘못된 입력:", e)
            return

    try:
        grid, start_cells, max_row, max_col = load_grid(filename)
    except Exception as e:
        print("엑셀 로드 중 오류:", e)
        return

    if len(start_cells) != 1:
        print("오류: START 셀은 정확히 1개 있어야 합니다.")
        return
    start = start_cells[0]
    print("START 지점:", grid[start]['coord'], "->", start)

    # 2) 경유지 입력 반복 (여러 경유지 입력 가능, q 입력 시 종료)
    while True:
        user_input = input("경유지(창고 or 복도; 소문자=창고, 대문자=복도, 여러 개는 쉼표로 구분) (q=종료): ").strip()
        if user_input.lower() == 'q':
            print("종료합니다.")
            break

        waypoint_strs = [x.strip() for x in user_input.split(",") if x.strip()]
        intermediate_nodes = []
        # warehouse_inputs 리스트: 창고로 입력된 값을 따로 기록
        warehouse_inputs = []
        for ws_val in waypoint_strs:
            if ws_val[0].islower():
                warehouse_inputs.append(ws_val)
                wh_cell = find_warehouse_cell(ws_val, grid)
                if not wh_cell:
                    print(f"창고 '{ws_val}' 셀이 존재하지 않습니다.")
                    continue
                corridor_cell = find_adjacent_corridor(wh_cell, grid)
                if not corridor_cell:
                    print(f"창고 '{ws_val}' 주변에 복도가 없습니다.")
                    continue
                print(f"창고 '{ws_val}' → 인접 복도 {grid[corridor_cell]['coord']}로 변환")
                intermediate_nodes.append(corridor_cell)
            else:
                node = coordinate_to_index(ws_val.upper())
                if node not in grid:
                    print(f"복도 '{ws_val}' 해당 셀이 존재하지 않습니다.")
                    continue
                intermediate_nodes.append(node)
        if not intermediate_nodes:
            print("유효한 경유지가 없습니다.")
            continue

        # 3) TSP 문제: START와 intermediate_nodes를 포함한 노드 집합 구성 (START는 첫 번째 노드)
        nodes = [start] + intermediate_nodes
        pairwise = compute_all_pairwise(nodes, grid)
        order, best_cost = solve_tsp(nodes, pairwise)
        if order is None:
            print("경유지를 모두 연결하는 경로를 찾을 수 없습니다.")
            continue

        print("최적 순서(노드 인덱스):", order, "총 거리:", best_cost)
        segments = []
        for i in range(len(order) - 1):
            _, seg = pairwise[(order[i], order[i+1])]
            if seg is None:
                print("경로 재구성에 실패했습니다.")
                segments = None
                break
            segments.append(seg)
        if segments is None:
            continue

        route_coords = [grid[p]['coord'] for seg in segments for p in seg]
        print("전체 경로 (셀 좌표):", " -> ".join(route_coords))
        tsp_nodes = nodes
        plot_grid_and_path_segments(grid, segments, max_row, max_col, tsp_nodes=tsp_nodes, warehouse_inputs=warehouse_inputs)

###############################################################################
# 실행
###############################################################################
if __name__ == "__main__":
    main()
